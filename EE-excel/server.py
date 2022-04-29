from flask import Flask
from flask import flash, redirect, render_template, request, session 
from flask import abort, url_for, send_file, send_from_directory

import os
from functools import wraps

from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl import load_workbook

import requests

import json

from EagleEye import * 

een_sessions = {}
 
app = Flask(__name__)


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return render_template('login.html')
        return f(*args, **kwargs)
    return decorated_function


@app.route('/excel/static/<path:filename>')
def send_static_files(filename):
    return send_from_directory('static/', filename)

 
@app.route('/excel/')
def home():
    if not session.get('logged_in'):
        return redirect('/excel/login')
    else:
        return render_template('index.html')
 
@app.route('/excel/login', methods=['GET', 'POST'])
def do_admin_login():
    if request.method == 'GET':
        if not session.get('logged_in'):
            return render_template('login.html')
        else:
            return redirect('/excel/')

    if request.method == 'POST':
        if 'id' in session:
            
            if session['id'] in een_sessions:
                een = een_sessions[session['id']]

            else:
                een = EagleEye()
                een_sessions[session['id']] = een 
        else:
            session['id'] = os.urandom(16)
            een = EagleEye()
            een_sessions[session['id']] = een

        
        if een.login(username=request.form['username'], password=request.form['password']):
            session['logged_in'] = True
            session['name'] = f"{een.user['first_name']} {een.user['last_name']}"    
            
        else:
            flash('wrong password!')
        
        return redirect('/excel/')
 
@app.route("/excel/logout")
def logout():
    session['logged_in'] = False
    session['name'] = ''
    
    if 'id' in session and session['id'] is not None:
        if een_sessions[session['id']]:
            del een_sessions[session['id']]
    
    session['id'] = None
    return home()
 

@app.route('/excel/generate', methods=['GET', 'POST'])
@login_required
def generate_data():
    if request.method == 'GET':
        een = een_sessions[session['id']]
        return render_template('devices.html', cameras=een.cameras)

    if request.method == 'POST':
        esn = request.form.get('esn')
        start_timestamp = request.form.get('start_time').replace('-','') + '000000.000' 
        end_timestamp = request.form.get('end_time').replace('-','') + '235959.999' 

        een = een_sessions[session['id']]
        camera = een.find_by_esn(esn)
        camera.get_video_list(instance=een, start_timestamp=start_timestamp, end_timestamp=end_timestamp, options='coalesce')

        wb = Workbook()
        ws = wb.active

        for video in camera.videos:
            ws.append([camera.camera_id, video[0], video[1]])

        filename = f"./uploads/{esn}-{start_timestamp}.xlsx"
     
        wb.save(filename)

        return send_file(filename, as_attachment=True)


@app.route('/excel/upload', methods=['POST'])
@login_required
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)

        f = request.files['file']
        location = secure_filename(f.filename)
        f.save('./uploads/' + location)
        return redirect('/excel/view/%s' % location, code=302)


@app.route('/excel/view/<path:filename>', methods=['GET'])
@login_required
def view(filename):
    wb = load_workbook('./uploads/' + filename)
    ret_obj = {
        'esns': {},
        'een': een_sessions[session['id']]
    }

    if wb:
        ws = wb.active

        for row in ws.iter_rows(row_offset=2):
            esn = row[0].value
            start = row[1].value
            end = row[2].value

            if esn:
                if esn in ret_obj['esns']:
                    ret_obj['esns'][esn].append(
                        {
                            'start_timestamp': start,
                            'end_timestamp': end
                        }   
                    )
                else:
                    ret_obj['esns'][esn] = [{
                        'start_timestamp': start,
                        'end_timestamp': end
                    }]

    return render_template('upload.html', template_values=ret_obj)


@app.route('/excel/api/get_info/<device_id>/<start_timestamp>/<end_timestamp>', methods=['GET'])
@login_required
def get_info(device_id, start_timestamp, end_timestamp):
    auth_key = een_sessions[session['id']].session.cookies['auth_key']
    url = f"https://login.eagleeyenetworks.com/asset/info/video?id={device_id}&start_timestamp={start_timestamp}&end_timestamp={end_timestamp}&A={auth_key}"
    print(url)
    print(auth_key)

    print('making get_info request')
    res = requests.get(url)
    print('finishing get_info request')

    print(f"get_info got a {res.status_code}")

    if res.status_code == 200:

        if res.text:
            return res.text


    return json.dumps({ 'proxy_status_code': res.status_code })


@app.route('/excel/api/fetch_video/<device_id>/<start_timestamp>/<end_timestamp>', methods=['GET'])
@login_required
def fetch_video(device_id, start_timestamp, end_timestamp):
    auth_key = een_sessions[session['id']].session.cookies['auth_key']
    url = f"{een_sessions[session['id']].host}/asset/cloud/video.flv?id={device_id}&start_timestamp={start_timestamp}&end_timestamp={end_timestamp}&webhook_url=https%3A%2F%2Feen.cloud%2Fdumpster%2F&A={auth_key}"
    print(url)
    print(auth_key)
    
    print('making fetch_video request')
    res = requests.get(url)
    print('finished fetch_video reqeust')
    
    print(f"fetch_video got a {res.status_code}")

    if res.status_code == 200:
        pass

    if res.status_code == 202:
        pass


    return json.dumps({ 'proxy_status_code': res.status_code })





if __name__ == "__main__":
    app.secret_key = os.urandom(12)
    app.run(debug=True,host='0.0.0.0', port=4000)
